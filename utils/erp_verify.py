# app/utils/erp_verify.py
from __future__ import annotations

import io
from typing import Any, Dict, List, Tuple
import openpyxl


def _s(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


def _to_float(v: Any) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except Exception:
        t = _s(v).replace(",", "")
        try:
            return float(t)
        except Exception:
            return 0.0


def parse_erp_excel_bytes(data: bytes) -> List[Dict[str, Any]]:
    """
    필수: 품번(또는 제품코드), 수량(또는 현재고)
    선택: LOT, 규격
    반환: [{item_code, lot, spec, qty}, ...]
    """
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    header = None
    for r in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header = list(r)
        break
    if not header:
        return []

    headers = {(_s(h).replace(" ", "")): idx for idx, h in enumerate(header) if _s(h)}

    code_keys = ["품번", "제품코드", "상품코드", "품목코드", "ITEMCODE", "CODE", "ITEM_CODE"]
    lot_keys = ["LOT", "로트", "LOTNO", "LOT번호", "LOTNO."]
    spec_keys = ["규격", "SPEC", "사이즈", "SIZE"]
    qty_keys = ["수량", "현재고", "재고", "QTY", "수량EA", "수량(EA)"]

    def pick(keys: List[str]):
        for k in keys:
            k2 = _s(k).replace(" ", "")
            if k2 in headers:
                return headers[k2]
        return None

    i_code = pick(code_keys)
    i_lot = pick(lot_keys)
    i_spec = pick(spec_keys)
    i_qty = pick(qty_keys)

    if i_code is None or i_qty is None:
        raise ValueError("엑셀 헤더에 '품번/제품코드' 또는 '수량/현재고' 컬럼이 없습니다.")

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = _s(row[i_code]) if i_code < len(row) else ""
        if not code:
            continue

        lot = _s(row[i_lot]) if (i_lot is not None and i_lot < len(row)) else ""
        spec = _s(row[i_spec]) if (i_spec is not None and i_spec < len(row)) else ""
        qty = _to_float(row[i_qty]) if i_qty < len(row) else 0.0

        if qty == 0:
            continue

        rows.append({"item_code": code, "lot": lot, "spec": spec, "qty": qty})

    return rows


def make_compare_key(item_code: str, lot: str, spec: str) -> Tuple[str, Tuple[str, ...]]:
    """
    1) 품번+LOT+규격
    2) 품번+LOT
    3) 품번+규격
    4) 품번
    """
    c = _s(item_code)
    l = _s(lot)
    s = _s(spec)

    if l and s:
        return "L3", (c, l, s)
    if l:
        return "L2_LOT", (c, l)
    if s:
        return "L2_SPEC", (c, s)
    return "L1", (c,)
