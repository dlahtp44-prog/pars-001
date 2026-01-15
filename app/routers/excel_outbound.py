from fastapi import APIRouter, UploadFile, File, Form, HTTPException
import openpyxl
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from app.db import query_inventory, upsert_inventory, add_history
from app.utils.excel_kor_columns import build_col_index

router = APIRouter(prefix="/api/excel/outbound", tags=["excel-outbound"])


# =====================================
# 🔥 수량 파싱 (소수점 유지)
# =====================================
def _parse_qty(v) -> float:
    if v is None:
        return 0.0

    s = str(v).strip()
    if s == "":
        return 0.0

    s = s.replace(",", "")

    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        raise ValueError("수량 형식 오류")


# =====================================
# 📅 엑셀 날짜 파싱 (출고일)
# =====================================
def _parse_excel_date(v):
    if v is None or str(v).strip() == "":
        return None

    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time())

    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d")
    except ValueError:
        raise ValueError("출고일 형식 오류 (YYYY-MM-DD)")


@router.post("")
async def excel_outbound(
    operator: str = Form(""),
    file: UploadFile = File(...)
):
    """
    출고 엑셀 업로드 (날짜 지정 지원)

    ✅ 필수 컬럼
      - 수량

    ⭕ 선택 컬럼
      - 출고일 (YYYY-MM-DD or 엑셀 날짜)
      - 창고
      - 로케이션
      - 브랜드
      - 품번
      - 품명
      - LOT
      - 규격
      - 비고

    📌 규칙
      - 수량 > 0 : 재고 차감 + 이력
      - 수량 = 0 : 재고 변화 없음 + 이력
      - 수량 < 0 : 에러
    """

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=400,
            detail="엑셀(.xlsx) 파일만 업로드 가능합니다."
        )

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_excel_outbound")

    data = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    # ===============================
    # HEADER
    # ===============================
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h for h in header_row]
    idx = build_col_index(headers)

    # 🔥 필수 컬럼: 수량만
    if "수량" not in idx:
        raise HTTPException(
            status_code=400,
            detail="필수 컬럼 누락: 수량"
        )

    success = 0
    fail = 0
    errors = []

    # ===============================
    # ROW LOOP
    # ===============================
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        try:
            # ---------------------------
            # 값 추출 (전부 선택)
            # ---------------------------
            warehouse = str(row[idx["창고"]] or "").strip() if "창고" in idx else ""
            location = str(row[idx["로케이션"]] or "").strip() if "로케이션" in idx else ""
            brand = str(row[idx["브랜드"]] or "").strip() if "브랜드" in idx else ""
            item_code = str(row[idx["품번"]] or "").strip() if "품번" in idx else ""
            item_name = str(row[idx["품명"]] or "").strip() if "품명" in idx else ""
            lot = str(row[idx["LOT"]] or "").strip() if "LOT" in idx else ""
            spec = str(row[idx["규격"]] or "").strip() if "규격" in idx else ""
            note = str(row[idx["비고"]] or "").strip() if "비고" in idx else ""

            qty = _parse_qty(row[idx["수량"]])

            # 📅 출고일 (선택)
            out_date = None
            if "출고일" in idx:
                out_date = _parse_excel_date(row[idx["출고일"]])

            if qty < 0:
                raise ValueError("수량은 0 이상만 허용")

            # ---------------------------
            # INVENTORY (qty > 0)
            # ---------------------------
            if qty > 0:
                rows = query_inventory(
                    warehouse=warehouse or None,
                    location=location or None,
                    brand=brand or None,
                    item_code=item_code or None,
                    lot=lot or None,
                    spec=spec or None,
                )

                if not rows:
                    raise ValueError("출고 가능한 재고가 없습니다.")

                remain = qty

                for r in rows:
                    if remain <= 0:
                        break

                    take = min(float(r["qty"]), remain)

                    ok = upsert_inventory(
                        r["warehouse"],
                        r["location"],
                        r["brand"],
                        r["item_code"],
                        r["item_name"],
                        r["lot"],
                        r["spec"],
                        -take,
                        note,
                    )
                    if not ok:
                        raise ValueError("재고 차감 실패")

                    add_history(
                        "출고",
                        r["warehouse"],
                        operator,
                        r["brand"],
                        r["item_code"],
                        r["item_name"],
                        r["lot"],
                        r["spec"],
                        r["location"],
                        "",
                        take,
                        note,
                        batch_id=batch_id,
                        created_at=out_date,   # 🔥 출고일 반영
                    )

                    remain -= take

                if remain > 0:
                    raise ValueError("출고 수량이 재고보다 많습니다.")

            else:
                # qty == 0 → 이력만 기록
                add_history(
                    "출고",
                    warehouse,
                    operator,
                    brand,
                    item_code,
                    item_name,
                    lot,
                    spec,
                    location,
                    "",
                    0,
                    note,
                    batch_id=batch_id,
                    created_at=out_date,
                )

            success += 1

        except Exception as e:
            fail += 1
            errors.append({
                "row": r_i,
                "error": str(e),
            })

    return {
        "ok": True,
        "success": success,
        "fail": fail,
        "batch_id": batch_id,
        "errors": errors[:50],
    }
