from fastapi import APIRouter, UploadFile, File, Form, HTTPException
import openpyxl
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from app.db import upsert_inventory, add_history
from app.utils.excel_kor_columns import build_col_index

router = APIRouter(prefix="/api/excel/inbound", tags=["excel-inbound"])


# =====================================
# 🔥 수량 파싱 (소수점 절대 보존)
# =====================================
def _parse_qty(v) -> float:
    if v is None:
        return 0.0

    s = str(v).strip()
    if s == "":
        return 0.0

    s = s.replace(",", "")

    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        raise ValueError("수량 형식 오류")


# =====================================
# 📅 엑셀 날짜 파싱 (입고일)
# =====================================
def _parse_excel_date(v):
    if v is None or str(v).strip() == "":
        return None

    # 엑셀 date / datetime 타입
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time())

    # 문자열 (YYYY-MM-DD)
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d")
    except ValueError:
        raise ValueError("입고일 형식 오류 (YYYY-MM-DD)")


@router.post("")
async def excel_inbound(
    operator: str = Form(""),
    file: UploadFile = File(...)
):
    """
    입고 엑셀 업로드

    ✅ 필수 컬럼
      - 수량

    ⭕ 선택 컬럼
      - 입고일 (YYYY-MM-DD or 엑셀 날짜)
      - 창고 / 로케이션 / 브랜드 / 품번 / 품명 / LOT / 규격 / 비고

    📌 규칙
      - 수량 > 0 : 재고 증가 + 이력
      - 수량 = 0 : 재고 변화 없음 + 이력
      - 수량 < 0 : 에러
    """

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx) 파일만 업로드 가능합니다.")

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_excel_inbound")

    data = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    # ===============================
    # HEADER
    # ===============================
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h for h in header_row]
    idx = build_col_index(headers)

    # 🔥 필수 컬럼: 수량만
    if "수량" not in idx:
        raise HTTPException(status_code=400, detail="필수 컬럼 누락: 수량")

    success = 0
    fail = 0
    errors = []

    # ===============================
    # ROW LOOP
    # ===============================
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        try:
            # ---------------------------
            # 값 추출 (전부 선택)
            # ---------------------------
            warehouse = str(row[idx["창고"]] or "").strip() if "창고" in idx else ""
            location = str(row[idx["로케이션"]] or "").strip() if "로케이션" in idx else ""
            item_code = str(row[idx["품번"]] or "").strip() if "품번" in idx else ""

            brand = str(row[idx["브랜드"]] or "").strip() if "브랜드" in idx else ""
            item_name = str(row[idx["품명"]] or "").strip() if "품명" in idx else ""
            lot = str(row[idx["LOT"]] or "").strip() if "LOT" in idx else ""
            spec = str(row[idx["규격"]] or "").strip() if "규격" in idx else ""
            note = str(row[idx["비고"]] or "").strip() if "비고" in idx else ""

            qty_raw = row[idx["수량"]]

            # 📅 입고일 (선택)
            in_date = None
            if "입고일" in idx:
                in_date = _parse_excel_date(row[idx["입고일"]])

            # ---------------------------
            # 수량 처리
            # ---------------------------
            qty = _parse_qty(qty_raw)
            if qty < 0:
                raise ValueError("수량은 0 이상만 허용")

            # ---------------------------
            # INVENTORY
            # ---------------------------
            if qty > 0:
                ok = upsert_inventory(
                    warehouse=warehouse,
                    location=location,
                    brand=brand,
                    item_code=item_code,
                    item_name=item_name,
                    lot=lot,
                    spec=spec,
                    qty_delta=qty,
                    note=note,
                )
                if not ok:
                    raise ValueError("재고 반영 실패")

            # ---------------------------
            # HISTORY (입고일 반영 🔥)
            # ---------------------------
            add_history(
                "입고",
                warehouse,
                operator,
                brand,
                item_code,
                item_name,
                lot,
                spec,
                "",
                location,
                qty,
                note,
                batch_id=batch_id,
                created_at=in_date,   # 🔥 핵심
            )

            success += 1

        except Exception as e:
            fail += 1
            errors.append({"row": r_i, "error": str(e)})

    return {
        "ok": True,
        "success": success,
        "fail": fail,
        "batch_id": batch_id,
        "errors": errors[:50],
    }
