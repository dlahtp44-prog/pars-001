# app/routers/api_init_inventory.py
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, List, Tuple

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from app.db import get_db, upsert_inventory, add_history

router = APIRouter(prefix="/api/init", tags=["초기재고 세팅"])

# =====================================================
# CONFIG
# =====================================================

REQUIRED_COLS = ["수량"]
OPTIONAL_COLS = ["창고", "로케이션", "브랜드", "품번", "품명", "LOT", "규격", "비고"]
ALL_COLS = REQUIRED_COLS + [c for c in OPTIONAL_COLS if c not in REQUIRED_COLS]


def _norm(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


def _q3(v: Any) -> Decimal:
    try:
        if v is None:
            raise ValueError
        s = str(v).strip()
        if s == "":
            raise ValueError
        return Decimal(s).quantize(Decimal("0.000"), rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0.000")


# =====================================================
# EXCEL PARSER (중복 키 → 수량 합산)
# =====================================================

def _read_excel_rows(data: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    try:
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        raise HTTPException(status_code=400, detail="엑셀에 데이터가 없습니다.")

    headers = [_norm(h) for h in header_cells]
    col_index = {h: i for i, h in enumerate(headers) if h}

    missing = [c for c in REQUIRED_COLS if c not in col_index]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"필수 컬럼 누락: {', '.join(missing)} (수량 컬럼 필수)",
        )

    ok_rows: List[Dict[str, Any]] = []
    err_rows: List[Dict[str, Any]] = []

    # -----------------------------
    # 1차 파싱
    # -----------------------------
    for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all((_norm(x) == "" for x in row)):
            continue

        def get(col: str) -> str:
            return _norm(row[col_index[col]]) if col in col_index else ""

        raw = {h: ("" if i >= len(row) else _norm(row[i])) for h, i in col_index.items()}

        qty = _q3(row[col_index["수량"]])
        if qty <= 0:
            err_rows.append({
                "rownum": ridx,
                "error": "수량은 0보다 커야 합니다.",
                "raw": raw
            })
            continue

        ok_rows.append({
            "warehouse": get("창고"),
            "location": get("로케이션"),
            "brand": get("브랜드"),
            "item_code": get("품번"),
            "item_name": get("품명"),
            "lot": get("LOT"),
            "spec": get("규격"),
            "qty": qty,
            "note": get("비고"),
        })

    # -----------------------------
    # 2차 처리: 중복 키 → 수량 합산
    # -----------------------------
    merged: Dict[
        Tuple[str, str, str, str, str, str],
        Dict[str, Any]
    ] = {}

    for r in ok_rows:
        key = (
            r["warehouse"],
            r["location"],
            r["brand"],
            r["item_code"],
            r["lot"],
            r["spec"],
        )

        if key not in merged:
            merged[key] = r.copy()
        else:
            merged[key]["qty"] = (
                merged[key]["qty"] + r["qty"]
            ).quantize(Decimal("0.000"), rounding=ROUND_HALF_UP)

    return list(merged.values()), err_rows


# =====================================================
# DB UTILS
# =====================================================

def _count_inventory() -> int:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM inventory")
        return int(cur.fetchone()[0])
    finally:
        conn.close()


def _count_history() -> int:
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM history")
        return int(cur.fetchone()[0])
    finally:
        conn.close()


def _make_batch_id() -> str:
    return "INIT-" + datetime.now().strftime("%Y%m%d-%H%M%S")


# =====================================================
# APIs
# =====================================================

@router.get("/status")
def init_inventory_status():
    return {
        "inventory_count": _count_inventory(),
        "history_count": _count_history(),
    }


@router.post("/preview")
async def init_preview(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx) 파일만 업로드 가능합니다.")

    data = await file.read()
    ok_rows, err_rows = _read_excel_rows(data)

    return {
        "ok": True,
        "summary": {
            "total_rows": len(ok_rows) + len(err_rows),
            "ok_rows": len(ok_rows),
            "error_rows": len(err_rows),
        },
        "rows_ok": ok_rows[:2000],
        "rows_error": err_rows[:2000],
        "message": f"정상 {len(ok_rows)}행 / 오류 {len(err_rows)}행 (중복은 자동 합산)",
    }


@router.post("/commit")
async def init_commit(
    file: UploadFile = File(...),
    operator: str = Form(...),
    confirm: str = Form(""),
    force: int = Form(0),
):
    if confirm.strip() != "INIT-CONFIRM":
        raise HTTPException(status_code=400, detail="confirm=INIT-CONFIRM 필요")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx) 파일만 업로드 가능합니다.")

    inv_cnt = _count_inventory()
    if inv_cnt > 0 and int(force) != 1:
        raise HTTPException(
            status_code=400,
            detail=f"inventory {inv_cnt}건 존재 → force=1 필요",
        )

    data = await file.read()
    ok_rows, err_rows = _read_excel_rows(data)

    if not ok_rows:
        raise HTTPException(status_code=400, detail="반영할 정상 데이터가 없습니다.")

    batch_id = _make_batch_id()
    applied = 0
    failed: List[Dict[str, Any]] = []

    for r in ok_rows:
        try:
            upsert_inventory(
                r["warehouse"],
                r["location"],
                r["brand"],
                r["item_code"],
                r["item_name"],
                r["lot"],
                r["spec"],
                float(r["qty"]),
                note=(r.get("note") or "초기재고"),
            )

            add_history(
                "초기재고",
                r["warehouse"],
                operator,
                r["brand"],
                r["item_code"],
                r["item_name"],
                r["lot"],
                r["spec"],
                "INIT",
                r["location"],
                float(r["qty"]),
                note="초기재고(엑셀 합산)",
                batch_id=batch_id,
                dedup_seconds=0,
            )

            applied += 1

        except Exception as e:
            failed.append({"row": r, "error": str(e)})

    return {
        "ok": True,
        "batch_id": batch_id,
        "summary": {
            "total_rows": len(ok_rows),
            "applied": applied,
            "failed": len(failed),
            "excel_errors": len(err_rows),
        },
        "failed_rows": failed[:200],
        "message": f"초기재고 반영 완료 (합산 기준)",
    }
