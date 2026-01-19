from fastapi import APIRouter, UploadFile, File, HTTPException, Request, Query
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates

from app.core.paths import TEMPLATES_DIR

import openpyxl
import qrcode
import base64
from io import BytesIO

router = APIRouter(prefix="/api/labels", tags=["라벨 API"])
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


# =====================================================
# 🏷️ 제품 라벨 (엑셀 업로드 → 미리보기)
# 규격: LS-3108 (99.1 × 38.1)
# =====================================================
@router.post("/product", response_class=HTMLResponse)
def product_label_preview(
    request: Request,
    file: UploadFile = File(...),
    spec: str = Query("3108")  # 제품 라벨
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="엑셀(xlsx) 파일만 업로드 가능합니다.")

    wb = openpyxl.load_workbook(file.file)
    ws = wb.active

    items = []

    # 엑셀 컬럼
    # A: 브랜드 / B: 품번 / C: 품명 / D: LOT / E: 규격
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue

        brand, code, name, lot, size = row

        brand = str(brand).strip()
        code = str(code).strip()
        name = str(name).strip()
        lot = str(lot).strip()
        size = str(size).strip()

        qr_text = f"PRODUCT:{code}|LOT:{lot}"

        qr = qrcode.make(qr_text)
        buffer = BytesIO()
        qr.save(buffer, format="PNG")
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()

        items.append({
            "brand": brand,
            "code": code,
            "name": name,
            "lot": lot,
            "spec": size,
            "qr_base64": qr_base64,
        })

    if not items:
        raise HTTPException(status_code=400, detail="출력할 제품 데이터가 없습니다.")

    return templates.TemplateResponse(
        "labels/product_preview.html",
        {
            "request": request,
            "items": items,
            "label_spec": spec,   # 반드시 전달
        }
    )


# =====================================================
# 📍 로케이션 라벨 (엑셀 업로드 → 미리보기)
# 규격: LS-3118 (99.1 × 140)
# =====================================================
@router.post("/location/excel", response_class=HTMLResponse)
def location_label_excel_preview(
    request: Request,
    file: UploadFile = File(...),
    spec: str = Query("3118")  # 로케이션 라벨
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="엑셀(xlsx) 파일만 업로드 가능합니다.")

    wb = openpyxl.load_workbook(file.file)
    ws = wb.active

    locations = []

    # 엑셀 컬럼
    # A: LOCATION
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        location = str(row[0]).strip().upper()
        qr_text = f"LOCATION:{location}"

        qr = qrcode.make(qr_text)
        buffer = BytesIO()
        qr.save(buffer, format="PNG")
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()

        locations.append({
            "location": location,
            "qr_base64": qr_base64
        })

    if not locations:
        raise HTTPException(status_code=400, detail="출력할 로케이션 데이터가 없습니다.")

    return templates.TemplateResponse(
        "labels/location_preview.html",
        {
            "request": request,
            "locations": locations,
            "label_spec": spec,   # 반드시 전달
        }
    )


# =====================================================
# 📍 로케이션 라벨 (단건 입력 → 미리보기)
# =====================================================
@router.get("/location", response_class=HTMLResponse)
def location_single_preview(
    request: Request,
    location: str = Query(...),
    spec: str = Query("3118")
):
    location = location.strip().upper()

    qr_text = f"LOCATION:{location}"
    qr = qrcode.make(qr_text)
    buffer = BytesIO()
    qr.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    return templates.TemplateResponse(
        "labels/location_preview.html",
        {
            "request": request,
            "locations": [{
                "location": location,
                "qr_base64": qr_base64
            }],
            "label_spec": spec,
        }
    )
