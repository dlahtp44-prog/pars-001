from __future__ import annotations

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
import openpyxl
import io
from datetime import datetime
from decimal import Decimal

from app.db import upsert_inventory, add_history
from app.utils.excel_kor_columns import build_col_index

router = APIRouter(prefix="/api/excel/inbound", tags=["excel-inbound"])


# =====================================
# 수량 파싱 (소수점 보존)
# =====================================
def _parse_qty(v) -> float:
    try:
        if v is None or str(v).strip() == "":
            return 0.0
        return float(Decimal(str(v)))
    except Exception:
        raise ValueError("수량 형식 오류")


@router.post("")
async def excel_inbound(
    operator: str = Form(""),
    file: UploadFile = File(...),
):
    """입고 엑셀 업로드 (한글 컬럼)

    필수 컬럼: 창고, 로케이션, 품번, 수량(0/빈값 허용)
    선택 컬럼: 브랜드, 품명, LOT, 규격, 비고

    규칙
    - 수량 > 0 : 재고 증가 + 이력
    - 수량 = 0 : 재고 변화 없음 + 이력
    - 수량 < 0 : 에러
    """

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx) 파일만 업로드 가능합니다.")

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_excel_inbound")

    data = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = build_col_index(headers)

    required_cols = ["창고", "로케이션", "품번", "수량"]
    missing = [c for c in required_cols if c not in idx]
    if missing:
        raise HTTPException(status_code=400, detail=f"필수 컬럼 누락: {', '.join(missing)}")

    success, fail = 0, 0
    errors = []

    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        try:
            warehouse = str(row[idx["창고"]] or "").strip()
            location = str(row[idx["로케이션"]] or "").strip()
            item_code = str(row[idx["품번"]] or "").strip()
            qty_raw = row[idx["수량"]]

            brand = str(row[idx["브랜드"]] or "").strip() if "브랜드" in idx else ""
            item_name = str(row[idx["품명"]] or "").strip() if "품명" in idx else ""
            lot = str(row[idx["LOT"]] or "").strip() if "LOT" in idx else ""
            spec = str(row[idx["규격"]] or "").strip() if "규격" in idx else ""
            note = str(row[idx["비고"]] or "").strip() if "비고" in idx else ""

            if not (warehouse and location and item_code):
                raise ValueError("필수 값(창고/로케이션/품번) 누락")

            qty = _parse_qty(qty_raw)
            if qty < 0:
                raise ValueError("수량은 0 이상만 허용")

            if qty > 0:
                ok = upsert_inventory(
                    warehouse=warehouse,
                    location=location,
                    brand=brand,
                    item_code=item_code,
                    item_name=item_name,
                    lot=lot,
                    spec=spec,
                    qty_delta=qty,
                    note=note,
                )
                if not ok:
                    raise ValueError("재고 반영 실패")

            add_history(
                "입고",
                warehouse,
                operator,
                brand,
                item_code,
                item_name,
                lot,
                spec,
                "",
                location,
                qty,
                note,
                batch_id=batch_id,
            )

            success += 1

        except Exception as e:
            fail += 1
            errors.append({"row": r_i, "error": str(e)})

    return {"ok": True, "success": success, "fail": fail, "batch_id": batch_id, "errors": errors[:50]}
