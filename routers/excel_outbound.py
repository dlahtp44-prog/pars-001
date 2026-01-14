from __future__ import annotations

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
import openpyxl
import io
from datetime import datetime
from decimal import Decimal

from app.db import upsert_inventory, add_history, query_inventory
from app.utils.excel_kor_columns import build_col_index, validate_required

router = APIRouter(prefix="/api/excel/outbound", tags=["excel-outbound"])


def _parse_qty(v) -> float:
    try:
        if v is None or str(v).strip() == "":
            raise ValueError("수량 누락")
        return float(Decimal(str(v)))
    except Exception:
        raise ValueError("수량 형식 오류")


def _resolve_when_lot_spec_missing(
    *,
    warehouse: str,
    location: str,
    brand: str,
    item_code: str,
    item_name: str,
):
    """LOT/규격이 비어있을 때, 현재고에서 후보를 찾아 1개로 확정"""
    rows = query_inventory(
        warehouse=warehouse or None,
        location=location or None,
        brand=brand or None,
        item_code=item_code or None,
        item_name=item_name or None,
        lot=None,
        spec=None,
        q="",
        limit=200,
    )
    if len(rows) == 1:
        r = rows[0]
        return r["brand"], r["item_name"], r["lot"], r["spec"]
    if len(rows) == 0:
        raise ValueError("현재고에서 후보를 찾을 수 없습니다. (로케이션/품번 확인)")
    raise ValueError(f"LOT/규격이 비어있고 후보가 {len(rows)}개입니다. (LOT/규격을 입력하세요)")


@router.post("")
async def excel_outbound(operator: str = Form(""), file: UploadFile = File(...)):
    """출고 엑셀 업로드 (한글 컬럼 고정)

    ✅ 필수: 로케이션, 품번, 품명, 수량
    ⭕ 선택: 창고, 브랜드, LOT, 규격, 비고

    - LOT/규격이 비어있으면 현재고에서 1개로 확정 가능한 경우에만 자동 매칭
    - 수량은 소수점 허용(REAL)
    """

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx) 파일만 업로드 가능합니다.")

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_excel_outbound")

    data = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = build_col_index(headers)

    required = ["로케이션", "품번", "품명", "수량"]
    ok, missing = validate_required(idx, required=required)
    if not ok:
        raise HTTPException(status_code=400, detail=f"필수 컬럼 누락: {', '.join(missing)}")

    success, fail = 0, 0
    errors = []

    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            warehouse = str(row[idx["창고"]] or "").strip() if "창고" in idx else ""
            brand = str(row[idx["브랜드"]] or "").strip() if "브랜드" in idx else ""

            location = str(row[idx["로케이션"]] or "").strip()
            item_code = str(row[idx["품번"]] or "").strip()
            item_name = str(row[idx["품명"]] or "").strip()

            lot = str(row[idx["LOT"]] or "").strip() if "LOT" in idx else ""
            spec = str(row[idx["규격"]] or "").strip() if "규격" in idx else ""

            qty_raw = row[idx["수량"]]
            note = str(row[idx["비고"]] or "").strip() if "비고" in idx else ""

            if not (location and item_code and item_name):
                raise ValueError("필수 값(로케이션/품번/품명) 누락")

            qty = _parse_qty(qty_raw)
            if qty <= 0:
                raise ValueError("수량은 0보다 커야 합니다.")

            # LOT/규격이 없으면 현재고에서 1개로 확정될 때만 자동 매칭
            if not lot or not spec:
                brand2, name2, lot2, spec2 = _resolve_when_lot_spec_missing(
                    warehouse=warehouse,
                    location=location,
                    brand=brand,
                    item_code=item_code,
                    item_name=item_name,
                )
                if not brand:
                    brand = brand2
                if not item_name:
                    item_name = name2
                lot = lot or lot2
                spec = spec or spec2

            # 재고 차감 (qty_delta 음수)
            ok2 = upsert_inventory(
                warehouse=warehouse,
                location=location,
                brand=brand,
                item_code=item_code,
                item_name=item_name,
                lot=lot,
                spec=spec,
                qty_delta=-qty,
                note=note,
            )
            if not ok2:
                raise ValueError("재고 반영 실패(재고 부족 또는 키 불일치)")

            add_history(
                "출고",
                warehouse,
                operator,
                brand,
                item_code,
                item_name,
                lot,
                spec,
                location,
                "",
                qty,
                note,
                batch_id=batch_id,
            )

            success += 1
        except Exception as e:
            fail += 1
            errors.append({"row": r_i, "error": str(e)})

    return {"ok": True, "success": success, "fail": fail, "batch_id": batch_id, "errors": errors[:50]}
